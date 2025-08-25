import pandas as pd
from pathlib import Path
from openpyxl import Workbook

def combine_special_and_csvs(special_file, csv_files, output="final.xlsx", sheet_name="Combined"):
    special = pd.read_excel(special_file, sheet_name=0)

    index_col = special.iloc[:, 0]
    special_data = special.iloc[:, 1:]  
    special_data.columns = pd.MultiIndex.from_tuples([("black_holes", c) for c in special_data.columns])

    dfs = [special_data]
    for f in csv_files:
        df = pd.read_csv(f).iloc[:, 3:]
        df.columns = pd.MultiIndex.from_tuples([(Path(f).stem, c) for c in df.columns])
        dfs.append(df)

    combined = pd.concat(dfs, axis=1)

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    ws.cell(row=1, column=1, value="Index")
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)

    catalogs_in_order = []
    for cat in combined.columns.get_level_values(0):
        if cat not in catalogs_in_order:
            catalogs_in_order.append(cat)

    # write header row 1 (catalog names) and row 2 (column names)
    col_idx = 2  
    for catalog in catalogs_in_order:
        cols = combined.loc[:, catalog]  
        if isinstance(cols, pd.Series):
            cols = cols.to_frame()
            cols.columns = pd.Index([cols.name[1]])

        ncols = len(cols.columns)
        ws.cell(row=1, column=col_idx, value=catalog)
        if ncols > 1:
            ws.merge_cells(start_row=1, start_column=col_idx, end_row=1, end_column=col_idx + ncols - 1)

        for i, col in enumerate(cols.columns, start=0):
            ws.cell(row=2, column=col_idx + i, value=str(col))

        col_idx += ncols

    # write the data (row 3+)
    for r in range(len(combined)):
        # index column
        ws.cell(row=r + 3, column=1, value=index_col.iloc[r])
        row_values = combined.iloc[r]
        for c, val in enumerate(row_values, start=2):
            ws.cell(row=r + 3, column=c, value=(val if pd.notna(val) else None))

    wb.save(output)

def main():
    first_file = "results/black_holes.xlsx"
    input_files = [
        "results/black_holes_with_2mass.csv",
        "results/black_holes_with_vvv_new.csv",
        "results/black_holes_with_UKIRT.csv",
        "results/black_holes_with_gaia.csv",
    ]
    combine_special_and_csvs(first_file, input_files, output="results/final.xlsx", sheet_name="Combined")

if __name__ == "__main__":
    main()
